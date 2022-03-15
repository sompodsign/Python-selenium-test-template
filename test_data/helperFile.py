import os
from openpyxl import load_workbook

filelist = os.listdir()
for file in filelist:
    if "xlsx" in file:
        wb = load_workbook(file)
        ws = wb.create_sheet()
        ws.title = "addCatalog"
        wb.save(file)
        wb.close()
        print("done"+file)
