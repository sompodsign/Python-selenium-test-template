import openpyxl
import pandas as pd


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rowno, colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rowno, colno).value


def writeData(file, sheetName, rowno, colno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rowno, colno).value = data
    workbook.save(file)


def read_data_from_excel(file, sheet_name):
    df = pd.read_excel(file, sheet_name)
    return df.to_dict(orient="records")[0]


def read_test_data_from_excel(file, sheet_name, table_name):
    df = pd.read_excel(file, sheet_name, table_name)
    print(df.to_dict(orient="records")[0])
    return df.to_dict(orient="records")[0]


def read_data_from_excel_by_row(file, sheet_name, row_value):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    data = []
    for index, row in enumerate(sheet.rows):
        if row[0].value == row_value:
            for cell in row:
                data.append(cell.value)
    return tuple(data[1:])
